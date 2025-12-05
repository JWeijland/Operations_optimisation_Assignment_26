#!/usr/bin/env python3
"""
Utility Functions for Aircraft Landing Scheduling Project
Helper functions for formatting, exporting, and analysis
"""

import json
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
from tabulate import tabulate

from .data_loader import ProblemInstance, Aircraft
from .model import Solution

# Try to import openpyxl for Excel export (optional)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Note: openpyxl not installed. Excel export will be skipped. Install with: pip install openpyxl")


def format_time(seconds: float) -> str:
    """
    Format time in human-readable format.

    Args:
        seconds: Time in seconds

    Returns:
        Formatted time string
    """
    if seconds < 1:
        return f"{seconds*1000:.1f}ms"
    elif seconds < 60:
        return f"{seconds:.2f}s"
    elif seconds < 3600:
        minutes = int(seconds // 60)
        secs = seconds % 60
        return f"{minutes}m {secs:.1f}s"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        return f"{hours}h {minutes}m"


def format_cost(cost: float) -> str:
    """
    Format cost with appropriate precision.

    Args:
        cost: Cost value

    Returns:
        Formatted cost string
    """
    if cost < 10:
        return f"{cost:.4f}"
    elif cost < 100:
        return f"{cost:.2f}"
    else:
        return f"{cost:.1f}"


def create_solution_summary(
    instance: ProblemInstance,
    solution: Solution
) -> Dict[str, Any]:
    """
    Create comprehensive summary of a solution.

    Args:
        instance: Problem instance
        solution: Solution to summarize

    Returns:
        Dictionary with summary statistics
    """
    # Calculate deviations
    early_count = 0
    late_count = 0
    on_time_count = 0
    total_early_time = 0.0
    total_late_time = 0.0
    max_early = 0.0
    max_late = 0.0

    for aircraft in instance.aircraft:
        landing_time = solution.get_landing_time(aircraft.id)
        deviation = landing_time - aircraft.target_time

        if abs(deviation) < 0.01:
            on_time_count += 1
        elif deviation < 0:
            early_count += 1
            total_early_time += abs(deviation)
            max_early = max(max_early, abs(deviation))
        else:
            late_count += 1
            total_late_time += deviation
            max_late = max(max_late, deviation)

    # Calculate runway utilization
    runway_counts = {}
    for runway in solution.runway_assignments.values():
        runway_counts[runway] = runway_counts.get(runway, 0) + 1

    summary = {
        'num_aircraft': instance.num_aircraft,
        'num_runways': len(set(solution.runway_assignments.values())),
        'total_cost': solution.objective_value,
        'solve_time': solution.solve_time,
        'status': solution.status,
        'early_count': early_count,
        'on_time_count': on_time_count,
        'late_count': late_count,
        'avg_early_time': total_early_time / early_count if early_count > 0 else 0,
        'avg_late_time': total_late_time / late_count if late_count > 0 else 0,
        'max_early_time': max_early,
        'max_late_time': max_late,
        'runway_utilization': runway_counts
    }

    return summary


def print_solution_details(
    instance: ProblemInstance,
    solution: Solution,
    verbose: bool = True
):
    """
    Print detailed solution information.

    Args:
        instance: Problem instance
        solution: Solution to print
        verbose: Include detailed aircraft information
    """
    print("\n" + "="*70)
    print("SOLUTION DETAILS")
    print("="*70)

    # Summary stats
    summary = create_solution_summary(instance, solution)

    print(f"\nGeneral Information:")
    print(f"  Aircraft:        {summary['num_aircraft']}")
    print(f"  Runways:         {summary['num_runways']}")
    print(f"  Total Cost:      {format_cost(summary['total_cost'])}")
    print(f"  Solve Time:      {format_time(summary['solve_time'])}")
    print(f"  Status:          {summary['status']}")

    print(f"\nLanding Statistics:")
    print(f"  On-time:         {summary['on_time_count']} aircraft")
    print(f"  Early:           {summary['early_count']} aircraft "
          f"(avg: {summary['avg_early_time']:.2f}, max: {summary['max_early_time']:.2f})")
    print(f"  Late:            {summary['late_count']} aircraft "
          f"(avg: {summary['avg_late_time']:.2f}, max: {summary['max_late_time']:.2f})")

    print(f"\nRunway Utilization:")
    for runway, count in sorted(summary['runway_utilization'].items()):
        percentage = (count / summary['num_aircraft']) * 100
        print(f"  Runway {runway}:      {count} aircraft ({percentage:.1f}%)")

    if verbose:
        print(f"\nDetailed Schedule:")
        print("-"*70)

        # Create table data
        table_data = []
        for aircraft in sorted(instance.aircraft, key=lambda a: a.id):
            landing_time = solution.get_landing_time(aircraft.id)
            runway = solution.get_runway(aircraft.id)
            deviation = landing_time - aircraft.target_time
            cost = aircraft.calculate_cost(landing_time)

            status = "On-time"
            if deviation < -0.01:
                status = f"Early {abs(deviation):.2f}"
            elif deviation > 0.01:
                status = f"Late {deviation:.2f}"

            table_data.append([
                f"A{aircraft.id}",
                runway,
                f"{aircraft.appearance_time:.1f}",
                f"{aircraft.target_time:.1f}",
                f"{aircraft.latest_time:.1f}",
                f"{landing_time:.1f}",
                status,
                f"{cost:.2f}"
            ])

        headers = ["Aircraft", "Runway", "E", "T", "L", "Landing", "Status", "Cost"]
        print(tabulate(table_data, headers=headers, tablefmt="grid"))

    print("="*70 + "\n")


def export_solution_json(
    instance: ProblemInstance,
    solution: Solution,
    filepath: str
):
    """
    Export solution to JSON format.

    Args:
        instance: Problem instance
        solution: Solution to export
        filepath: Output file path
    """
    # Create export data
    data = {
        'metadata': {
            'num_aircraft': instance.num_aircraft,
            'num_runways': len(set(solution.runway_assignments.values())),
            'total_cost': solution.objective_value,
            'solve_time': solution.solve_time,
            'status': solution.status
        },
        'aircraft': []
    }

    for aircraft in instance.aircraft:
        landing_time = solution.get_landing_time(aircraft.id)
        runway = solution.get_runway(aircraft.id)

        aircraft_data = {
            'id': aircraft.id,
            'appearance_time': aircraft.appearance_time,
            'target_time': aircraft.target_time,
            'latest_time': aircraft.latest_time,
            'landing_time': landing_time,
            'runway': runway,
            'deviation': landing_time - aircraft.target_time,
            'cost': aircraft.calculate_cost(landing_time)
        }
        data['aircraft'].append(aircraft_data)

    # Write to file
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    with open(filepath, 'w') as f:
        json.dump(data, f, indent=2)

    print(f"Solution exported to: {filepath}")


def export_detailed_solution_table(
    instance: ProblemInstance,
    solution: Solution,
    filepath: str,
    solution_type: str = "Solution"
):
    """
    Export detailed solution table to CSV with all aircraft information.

    Args:
        instance: Problem instance
        solution: Solution to export
        filepath: Output file path (CSV)
        solution_type: Type of solution (e.g., "Heuristic", "Optimal")
    """
    # Prepare data rows
    rows = []

    for aircraft in instance.aircraft:
        landing_time = solution.get_landing_time(aircraft.id)
        runway = solution.get_runway(aircraft.id)
        deviation = landing_time - aircraft.target_time
        cost = aircraft.calculate_cost(landing_time)

        # Determine if early or late
        if landing_time < aircraft.target_time:
            status = "Early"
            penalty_used = aircraft.early_penalty
            time_diff = aircraft.target_time - landing_time
        elif landing_time > aircraft.target_time:
            status = "Late"
            penalty_used = aircraft.late_penalty
            time_diff = landing_time - aircraft.target_time
        else:
            status = "On-time"
            penalty_used = 0
            time_diff = 0

        row = {
            'Aircraft_ID': f'A{aircraft.id}',
            'Runway': runway,
            'Earliest_Time': aircraft.appearance_time,
            'Target_Time': aircraft.target_time,
            'Latest_Time': aircraft.latest_time,
            'Actual_Landing_Time': landing_time,
            'Deviation': deviation,
            'Status': status,
            'Time_Difference': time_diff,
            'Early_Penalty_Rate': aircraft.early_penalty,
            'Late_Penalty_Rate': aircraft.late_penalty,
            'Penalty_Applied': penalty_used,
            'Cost': cost
        }
        rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Sort by runway and landing time
    df = df.sort_values(['Runway', 'Actual_Landing_Time'])

    # Add summary row
    summary_row = {
        'Aircraft_ID': 'TOTAL',
        'Runway': '-',
        'Earliest_Time': '-',
        'Target_Time': '-',
        'Latest_Time': '-',
        'Actual_Landing_Time': '-',
        'Deviation': '-',
        'Status': '-',
        'Time_Difference': '-',
        'Early_Penalty_Rate': '-',
        'Late_Penalty_Rate': '-',
        'Penalty_Applied': '-',
        'Cost': df['Cost'].sum()
    }
    df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)

    # Write to CSV
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    df.to_csv(filepath, index=False, float_format='%.2f')

    print(f"Detailed solution table exported to: {filepath}")

    return df


def export_detailed_solution_excel(
    instance: ProblemInstance,
    solution: Solution,
    filepath: str,
    solution_type: str = "Solution"
):
    """
    Export detailed solution table to Excel with beautiful formatting.

    Args:
        instance: Problem instance
        solution: Solution to export
        filepath: Output file path (Excel .xlsx)
        solution_type: Type of solution (e.g., "Heuristic", "Optimal")
    """
    # Check if openpyxl is available
    if not EXCEL_AVAILABLE:
        print(f"⚠ Skipping Excel export to {filepath} - openpyxl not installed")
        print("  Install with: pip install openpyxl")
        return

    # Prepare data rows
    rows = []

    for aircraft in instance.aircraft:
        landing_time = solution.get_landing_time(aircraft.id)
        runway = solution.get_runway(aircraft.id)
        deviation = landing_time - aircraft.target_time
        cost = aircraft.calculate_cost(landing_time)

        # Determine if early or late
        if landing_time < aircraft.target_time:
            status = "Early"
            penalty_used = aircraft.early_penalty
            time_diff = aircraft.target_time - landing_time
        elif landing_time > aircraft.target_time:
            status = "Late"
            penalty_used = aircraft.late_penalty
            time_diff = landing_time - aircraft.target_time
        else:
            status = "On-time"
            penalty_used = 0
            time_diff = 0

        row = {
            'Aircraft ID': f'A{aircraft.id}',
            'Runway': runway,
            'Earliest Time (min)': aircraft.appearance_time,
            'Target Time (min)': aircraft.target_time,
            'Latest Time (min)': aircraft.latest_time,
            'Actual Landing (min)': landing_time,
            'Deviation (min)': deviation,
            'Status': status,
            'Time Diff (min)': time_diff,
            'Early Penalty (€/min)': aircraft.early_penalty,
            'Late Penalty (€/min)': aircraft.late_penalty,
            'Penalty Applied (€/min)': penalty_used,
            'Cost (€)': cost
        }
        rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Sort by runway and landing time
    df = df.sort_values(['Runway', 'Actual Landing (min)'])

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{solution_type} Solution"

    # Add title
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f"Aircraft Landing Schedule - {solution_type} Solution"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Add metadata
    ws['A2'] = f"Total Aircraft: {instance.num_aircraft}"
    ws['A3'] = f"Total Cost: €{solution.objective_value:.2f}"
    ws['A4'] = f"Solve Time: {solution.solve_time:.3f}s"
    ws['A5'] = f"Status: {solution.status}"

    for row in range(2, 6):
        ws[f'A{row}'].font = Font(bold=True)

    # Write headers (row 7)
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Write data
    for row_num, row_data in enumerate(df.values, 8):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format numbers
            if col_num in [3, 4, 5, 6, 7, 9]:  # Time columns
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
            elif col_num in [10, 11, 12, 13]:  # Money columns
                if isinstance(value, (int, float)):
                    cell.number_format = '€#,##0.00'

            # Color code status
            if col_num == 8:  # Status column
                if value == "Early":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif value == "Late":
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif value == "On-time":
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Add TOTAL row
    total_row = len(df) + 8
    ws.cell(total_row, 1).value = "TOTAL"
    ws.cell(total_row, 13).value = df['Cost (€)'].sum()
    ws.cell(total_row, 13).number_format = '€#,##0.00'

    for col in range(1, 14):
        cell = ws.cell(total_row, col)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

    # Adjust column widths
    column_widths = {
        'A': 12, 'B': 8, 'C': 18, 'D': 18, 'E': 18, 'F': 18,
        'G': 16, 'H': 10, 'I': 16, 'J': 20, 'K': 20, 'L': 22, 'M': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save file
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb.save(filepath)

    print(f"Detailed Excel table exported to: {filepath}")

    return df


def create_latex_table(
    df: pd.DataFrame,
    caption: str = "",
    label: str = ""
) -> str:
    """
    Convert DataFrame to LaTeX table format.

    Args:
        df: DataFrame to convert
        caption: Table caption
        label: LaTeX label for referencing

    Returns:
        LaTeX table string
    """
    # Format numeric columns
    df_formatted = df.copy()
    for col in df_formatted.columns:
        if df_formatted[col].dtype in [np.float64, np.float32]:
            if 'Time' in col:
                df_formatted[col] = df_formatted[col].apply(lambda x: f"{x:.4f}")
            elif 'Cost' in col:
                df_formatted[col] = df_formatted[col].apply(lambda x: f"{x:.2f}")
            elif 'Gap' in col or '%' in col:
                df_formatted[col] = df_formatted[col].apply(lambda x: f"{x:.2f}")
            else:
                df_formatted[col] = df_formatted[col].apply(lambda x: f"{x:.2f}")

    # Generate LaTeX
    latex = df_formatted.to_latex(
        index=False,
        escape=False,
        column_format='l' + 'r' * (len(df.columns) - 1)
    )

    # Add caption and label
    if caption:
        latex = latex.replace(
            r'\end{tabular}',
            f'\\end{{tabular}}\n\\caption{{{caption}}}'
        )
    if label:
        latex = latex.replace(
            r'\end{table}',
            f'\\label{{{label}}}\n\\end{{table}}'
        )

    return latex


def calculate_statistics(values: List[float]) -> Dict[str, float]:
    """
    Calculate summary statistics for a list of values.

    Args:
        values: List of numeric values

    Returns:
        Dictionary with statistics
    """
    if not values:
        return {
            'mean': 0, 'median': 0, 'std': 0,
            'min': 0, 'max': 0, 'range': 0
        }

    return {
        'mean': np.mean(values),
        'median': np.median(values),
        'std': np.std(values),
        'min': np.min(values),
        'max': np.max(values),
        'range': np.max(values) - np.min(values)
    }


def print_table(
    data: List[List[Any]],
    headers: List[str],
    title: Optional[str] = None,
    tablefmt: str = "grid"
):
    """
    Print formatted ASCII table.

    Args:
        data: Table data as list of rows
        headers: Column headers
        title: Optional table title
        tablefmt: Table format (grid, simple, etc.)
    """
    if title:
        print(f"\n{title}")
        print("-" * len(title))

    print(tabulate(data, headers=headers, tablefmt=tablefmt))
    print()


def validate_solution(
    instance: ProblemInstance,
    solution: Solution,
    tolerance: float = 1e-6
) -> Tuple[bool, List[str]]:
    """
    Validate solution for feasibility.

    Args:
        instance: Problem instance
        solution: Solution to validate
        tolerance: Numerical tolerance

    Returns:
        Tuple of (is_valid, list_of_violations)
    """
    violations = []

    # Check time windows
    for aircraft in instance.aircraft:
        landing_time = solution.get_landing_time(aircraft.id)

        if landing_time < aircraft.appearance_time - tolerance:
            violations.append(
                f"Aircraft {aircraft.id}: Lands before appearance time "
                f"({landing_time:.2f} < {aircraft.appearance_time:.2f})"
            )

        if landing_time > aircraft.latest_time + tolerance:
            violations.append(
                f"Aircraft {aircraft.id}: Lands after latest time "
                f"({landing_time:.2f} > {aircraft.latest_time:.2f})"
            )

    # Check separations
    schedule_by_runway = solution.get_schedule_by_runway()

    for runway, landings in schedule_by_runway.items():
        for i in range(len(landings) - 1):
            aircraft_i_id, time_i = landings[i]
            aircraft_j_id, time_j = landings[i + 1]

            # Get indices
            idx_i = next(idx for idx, a in enumerate(instance.aircraft)
                        if a.id == aircraft_i_id)
            idx_j = next(idx for idx, a in enumerate(instance.aircraft)
                        if a.id == aircraft_j_id)

            required_sep = instance.get_separation(idx_i, idx_j)
            actual_sep = time_j - time_i

            if actual_sep < required_sep - tolerance:
                violations.append(
                    f"Runway {runway}: Insufficient separation between "
                    f"A{aircraft_i_id} and A{aircraft_j_id} "
                    f"({actual_sep:.2f} < {required_sep:.2f})"
                )

    is_valid = len(violations) == 0
    return is_valid, violations


def compare_solutions(
    instance: ProblemInstance,
    solutions: Dict[str, Solution],
    metric: str = 'cost'
) -> pd.DataFrame:
    """
    Compare multiple solutions.

    Args:
        instance: Problem instance
        solutions: Dictionary mapping name to Solution
        metric: Comparison metric ('cost', 'time', 'both')

    Returns:
        DataFrame with comparison
    """
    data = []

    for name, solution in solutions.items():
        row = {
            'Method': name,
            'Cost': solution.objective_value,
            'Time (s)': solution.solve_time,
            'Status': solution.status
        }

        # Add summary statistics
        summary = create_solution_summary(instance, solution)
        row['Early'] = summary['early_count']
        row['On-time'] = summary['on_time_count']
        row['Late'] = summary['late_count']

        data.append(row)

    df = pd.DataFrame(data)

    # Sort by metric
    if metric == 'cost':
        df = df.sort_values('Cost')
    elif metric == 'time':
        df = df.sort_values('Time (s)')

    return df
